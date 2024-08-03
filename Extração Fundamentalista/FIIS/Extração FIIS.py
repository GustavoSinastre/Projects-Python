# Importing the libraries used in this project
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options


# Importing the sheet's useds
workbook_path = 'Extração Fundamentalista/FIIS/Lista de FIIS.xlsx'
workbook = openpyxl.load_workbook(workbook_path)
search_page = workbook['Pesquisa']
final_page = workbook['Resultado']

# Clear the final page starting from the second row
final_page.delete_rows(1, final_page.max_row)

# Check if there are data in the search page
if search_page.max_row < 2:
    print("No data to process in the search page.")
else:
    # Set up Chrome options
    chrome_options = Options()
    #chrome_options.add_argument("--headless")  # Uncomment if you want to run Chrome headless

    # Initialize the WebDriver with options
    driver = webdriver.Chrome(options=chrome_options)
    wait = WebDriverWait(driver, 10)  # Adjusted wait time

    # Creating the search loop
    headers_written = False

    for row in search_page.iter_rows(min_row=2, values_only=True):
        ATIVO, TIPO = row
        # Open the website with the complete link
        url_path = 'https://investidor10.com.br/'
        complete_url = f"{url_path}{TIPO}{ATIVO}"  # Ensure the URL is properly formatted
        driver.get(complete_url)

        try:
            headers = []
            values = []

            # INFOS VALOR PATRIMONIAL
            for i in range(2, 4):
                try:
                    header_xpath = f"//*[@id='asset-value-comp']/div/div[{i}]/h4"
                    header_element = wait.until(EC.visibility_of_element_located((By.XPATH, header_xpath)))
                    header = header_element.text
                    headers.append(header)

                    value_xpath = f"//*[@id='asset-value-comp']/div/div[{i}]/div/div[1]"
                    value_element = wait.until(EC.visibility_of_element_located((By.XPATH, value_xpath)))
                    value = value_element.text
                    values.append(value)
                except Exception as e:
                    print(f"Error getting INFOS VALOR PATRIMONIAL data at index {i}: {e}")
                    headers.append(f"Header {i}")
                    values.append("-")
            for i in range(1, 4):
                try:
                    header_xpath = f"//*[@id='asset-value-comp']/div/div[4]/div[{i}]/span[1]"
                    header_element = wait.until(EC.visibility_of_element_located((By.XPATH, header_xpath)))
                    header = header_element.text
                    headers.append(header)

                    value_xpath = f"//*[@id='asset-value-comp']/div/div[4]/div[{i}]/span[2]"
                    value_element = wait.until(EC.visibility_of_element_located((By.XPATH, value_xpath)))
                    value = value_element.text
                    values.append(value)
                except Exception as e:
                    print(f"Error getting INFOS VALOR PATRIMONIAL data at index {i}: {e}")
                    headers.append(f"Header {i}")
                    values.append("-")

            # INFORMAÇÕES ADMINISTRATIVAS
            for i in range(1, 16):
                try:
                    header_xpath = f"//*[@id='table-indicators']/div[{i}]/div[2]/span"
                    header_element = wait.until(EC.visibility_of_element_located((By.XPATH, header_xpath)))
                    header = header_element.text
                    headers.append(header)

                    value_xpath = f"//*[@id='table-indicators']/div[{i}]/div[2]/div/span"
                    value_element = wait.until(EC.visibility_of_element_located((By.XPATH, value_xpath)))
                    value = value_element.text
                    values.append(value)
                except Exception as e:
                    print(f"Error getting INFORMAÇÕES ADMINISTRATIVAS data at index {i}: {e}")
                    headers.append(f"Header {i}")
                    values.append("-")

            # RENTABILIDADE
            for i in range(2, 8):
                try:
                    header_xpath = f"//*[@id='busca-avancada']/section/div/div[6]/div/div/div[{i}]/h4"
                    header_element = wait.until(EC.visibility_of_element_located((By.XPATH, header_xpath)))
                    header = header_element.text
                    headers.append(header)

                    value_xpath = f"//*[@id='busca-avancada']/section/div/div[6]/div/div/div[{i}]/span"
                    value_element = wait.until(EC.visibility_of_element_located((By.XPATH, value_xpath)))
                    value = value_element.text
                    values.append(value)
                except Exception as e:
                    print(f"Error getting INDICADORES FUNDAMENTALISTAS data at index {i}: {e}")
                    headers.append(f"Header {i}")
                    values.append("-")
            # RENTABILIDADE REAL
            for i in range(10, 15):
                try:
                    header_xpath = f"//*[@id='busca-avancada']/section/div/div[6]/div/div/div[{i}]/h4"
                    header_element = wait.until(EC.visibility_of_element_located((By.XPATH, header_xpath)))
                    header = header_element.text
                    headers.append(header)

                    value_xpath = f"//*[@id='busca-avancada']/section/div/div[6]/div/div/div[{i}]/span"
                    value_element = wait.until(EC.visibility_of_element_located((By.XPATH, value_xpath)))
                    value = value_element.text
                    values.append(value)
                except Exception as e:
                    print(f"Error getting INDICADORES FUNDAMENTALISTAS data at index {i}: {e}")
                    headers.append(f"Header {i}")
                    values.append("-")

            # DISTRIBUIÇÃO DE DIVIDENDOS    
            for i in range(1, 5):
                try:
                    header_xpath = f"//*[@id='yield-distribuition']/div/div[1]/div[{i}]/span[1]"
                    header_element = wait.until(EC.visibility_of_element_located((By.XPATH, header_xpath)))
                    header = header_element.text
                    headers.append(header)

                    value_xpath = f"//*[@id='yield-distribuition']/div/div[1]/div[{i}]/span[2]"
                    value_element = wait.until(EC.visibility_of_element_located((By.XPATH, value_xpath)))
                    value = value_element.text
                    values.append(value)
                except Exception as e:
                    print(f"Error getting DISTRIBUIÇÃO DE DIVIDENDOS data at index {i}: {e}")
                    headers.append(f"Header {i}")
                    values.append("-")

            # DIVIDENDOS DOS ÚLTIMOS 5 ANOS
            try:
                header_xpath = "//*[@id='dividend-yield-section']/div/div[2]/h3[2]"
                header_element = wait.until(EC.visibility_of_element_located((By.XPATH, header_xpath)))
                header = header_element.text
                headers.append(header)

                value_xpath = "//*[@id='dividend-yield-section']/div/div[2]/h3[2]/span"
                value_element = wait.until(EC.visibility_of_element_located((By.XPATH, value_xpath)))
                value = value_element.text
                values.append(value)
            except Exception as e:
                print(f"Error getting DIVIDENDOS DOS ÚLTIMOS 5 ANOS data: {e}")
                headers.append("DIVIDENDOS DOS ÚLTIMOS 5 ANOS")
                values.append("-")

            # Add data to final page
            if not headers_written:
                final_page.append(['ATIVO', 'TIPO'] + headers)
                headers_written = True

            final_page.append([ATIVO, TIPO] + values)
            print(f"Values written: {values}")

        except Exception as e:
            print(f"Error for {ATIVO} - {TIPO}: {e}")

    # Save changes to workbook
    workbook.save(workbook_path)
    print("Workbook saved.")

    # Close the browser
    driver.quit()
    print("Driver closed.")
