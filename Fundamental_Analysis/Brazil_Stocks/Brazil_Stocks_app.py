# Importing the libraries used in this project
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options

# Importing the sheet's used
workbook_path = r'C:\Users\luisg\OneDrive\Project Dev\GitHub\Python-Projects\Fundamental_Analysis\Brazil_Stocks\Brazil_Stocks_List.xlsx'
workbook = openpyxl.load_workbook(workbook_path)
search_page = workbook['Pesquisa']
final_page = workbook['Resultado']

# Clear the final page starting from the second row
final_page.delete_rows(1, final_page.max_row)

# Define fixed headers
fixed_headers = [
    'ATIVO', 'TIPO', 'COTAÇÃO', 'VARIAÇÃO (12M)', 'Rent. 1 Mês', 'Rent. 3 Meses', 'Rent. 1 Ano', 'Rent. 2 Anos', 'Rent. 5 Anos', 'Rent. 10 Anos', 'Rent. Real 3 Meses', 'Rent. Real 1 Ano', 'Rent. Real 2 Anos', 'Rent. Real 5 Anos', 'Rent. Real 10 Anos', 'P/L', 'P/RECEITA (PSR)', 'P/VP', 'DIVIDEND YIELD - CIEL3', 'PAYOUT', 'MARGEM LÍQUIDA', 'MARGEM BRUTA', 'MARGEM EBIT', 'MARGEM EBITDA', 'EV/EBITDA', 'EV/EBIT', 'P/EBITDA', 
    'P/EBIT', 'P/ATIVO', 'P/CAP.GIRO', 'P/ATIVO CIRC LIQ', 'VPA', 'LPA', 'GIRO ATIVOS', 'ROE', 'ROIC', 'ROA', 
    'DÍVIDA LÍQUIDA / PATRIMÔNIO', 'DÍVIDA LÍQUIDA / EBITDA', 'DÍVIDA LÍQUIDA / EBIT', 'DÍVIDA BRUTA / PATRIMÔNIO', 
    'PATRIMÔNIO / ATIVOS', 'PASSIVOS / ATIVOS', 'LIQUIDEZ CORRENTE', 'CAGR RECEITAS 5 ANOS', 'CAGR LUCROS 5 ANOS', 
    'DY médio em 5 anos: 5,26%', 'Nome da Empresa:', 'CNPJ:', 'Ano de estreia na bolsa:', 'Número de funcionários:', 
    'Ano de fundação:', 'Valor de mercado', 'Valor de firma', 'Patrimônio Líquido', 'Nº total de papeis', 'Ativos', 
    'Ativo Circulante', 'Dívida Bruta', 'Dívida Líquida', 'Disponibilidade'
]

# Add headers to final page if not already done
final_page.append(fixed_headers)

# Check if there are data in the search page
if search_page.max_row < 2:
    print("No data to process in the search page.")
else:
    # Set up Chrome options
    chrome_options = Options()
    #chrome_options.add_argument("--headless")  # Uncomment if you want to run Chrome headless

    # Initialize the WebDriver with options
    driver = webdriver.Chrome(options=chrome_options)
    wait = WebDriverWait(driver, 10)  # Adjusted wait time

    # Creating the search loop
    for row in search_page.iter_rows(min_row=2, values_only=True):
        ATIVO, TIPO = row
        # Open the website with the complete link
        url_path = 'https://investidor10.com.br/'
        complete_url = f"{url_path}{TIPO}{ATIVO}"  # Ensure the URL is properly formatted
        driver.get(complete_url)

        try:
            # Check for "404" or "Not Found" in the specific XPaths
            not_found_xpath1 = "/html/body/div/div[1]"
            not_found_xpath2 = "/html/body/div/div[2]"

            not_found_text1 = ""
            not_found_text2 = ""

            try:
                not_found_element1 = wait.until(EC.visibility_of_element_located((By.XPATH, not_found_xpath1)))
                not_found_text1 = not_found_element1.text
            except Exception:
                pass  # Ignore exceptions in this block, just move to the next

            try:
                not_found_element2 = wait.until(EC.visibility_of_element_located((By.XPATH, not_found_xpath2)))
                not_found_text2 = not_found_element2.text
            except Exception:
                pass  # Ignore exceptions in this block, just move to the next

            if "404" in not_found_text1 or "Not Found" in not_found_text1 or "404" in not_found_text2 or "Not Found" in not_found_text2:
                print(f"Page not found for {ATIVO} - {TIPO}. Skipping to the next.")
                continue

            # Initialize values list after checking for "404" or "Not Found"
            values = []

            # COTAÇÃO E VARIAÇÃO 12M
            for i in range(1, 3):
                try:
                    value_xpath = f"//*[@id='cards-ticker']/div[{i}]/div[2]/div/span"
                    value_element = wait.until(EC.visibility_of_element_located((By.XPATH, value_xpath)))
                    value = value_element.text
                    values.append(value)
                except Exception as e:
                    print(f"Error getting COTAÇÃO E VARIAÇÃO 12M data at index {i}: {e}")
                    values.append("-")

            # RENTABILIDADE DETALHADA
            for i in range(2, 8):
                try:
                    value_xpath = f"//*[@id='ticker']/section/div/div[2]/div/div/div[{i}]/span"
                    value_element = wait.until(EC.visibility_of_element_located((By.XPATH, value_xpath)))
                    value = value_element.text
                    values.append(value)
                except Exception as e:
                    print(f"Error getting RENTABILIDADE DETALHADA data at index {i}: {e}")
                    values.append("-")

            # RENTABILIDADE REAL DETALHADA
            for i in range(10, 15):
                try:
                    value_xpath = f"//*[@id='ticker']/section/div/div[2]/div/div/div[{i}]/span"
                    value_element = wait.until(EC.visibility_of_element_located((By.XPATH, value_xpath)))
                    value = value_element.text
                    values.append(value)
                except Exception as e:
                    print(f"Error getting RENTABILIDADE REAL DETALHADA data at index {i}: {e}")
                    values.append("-")

            # INDICADORES FUNDAMENTALISTAS
            for i in range(1, 32):
                try:
                    value_xpath = f"//*[@id='table-indicators']/div[{i}]/div[1]/span"
                    value_element = wait.until(EC.visibility_of_element_located((By.XPATH, value_xpath)))
                    value = value_element.text
                    values.append(value)
                except Exception as e:
                    print(f"Error getting INDICADORES FUNDAMENTALISTAS data at index {i}: {e}")
                    values.append("-")

            # DIVIDENDOS DOS ÚLTIMOS 5 ANOS
            try:
                value_xpath = "//*[@id='dividends-section']/div[1]/div[1]/h3[2]/span"
                value_element = wait.until(EC.visibility_of_element_located((By.XPATH, value_xpath)))
                value = value_element.text
                values.append(value)
            except Exception as e:
                print(f"Error getting DIVIDENDOS DOS ÚLTIMOS 5 ANOS data: {e}")
                values.append("-")

            # DADOS SOBRE A EMPRESA
            for i in range(1, 6):
                try:
                    value_xpath = f"//*[@id='data_about']/div[2]/div/div[1]/table/tbody/tr[{i}]/td[2]"
                    value_element = wait.until(EC.visibility_of_element_located((By.XPATH, value_xpath)))
                    value = value_element.text
                    values.append(value)
                except Exception as e:
                    print(f"Error getting DADOS SOBRE A EMPRESA data at index {i}: {e}")
                    values.append("-")

            # INFORMAÇÕES SOBRE A EMPRESA
            for i in range(1, 10):
                try:   
                    value_xpath = f"//*[@id='table-indicators-company']/div[{i}]/span[2]/div[1]"
                    value_element = wait.until(EC.visibility_of_element_located((By.XPATH, value_xpath)))
                    value = value_element.text
                    values.append(value)
                except Exception as e:
                    print(f"Error getting INFORMAÇÕES SOBRE A EMPRESA data at index {i}: {e}")
                    values.append("-")

            # Add data to final page
            final_page.append([ATIVO, TIPO] + values)
            print(f"Values written: {values}")

        except Exception as e:
            print(f"Error for {ATIVO} - {TIPO}: {e}")

    # Save changes to workbook
    workbook.save(workbook_path)
    print("Workbook saved.")

    # Close the browser
    driver.quit()
    print("Driver closed.")
